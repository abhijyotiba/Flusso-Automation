"""
Enhanced Multimodal Attachment Processor
Intelligent hybrid system for processing all attachment types with cost optimization
"""

import logging
import io
import time
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass
from enum import Enum
import requests
from requests.auth import HTTPBasicAuth
from google import genai

from app.config.settings import settings

logger = logging.getLogger(__name__)


class AttachmentType(Enum):
    """Classification of attachment types for routing"""
    TEXT_PDF = "text_pdf"              # Native text PDF (PyMuPDF)
    SCANNED_PDF = "scanned_pdf"        # Scanned/image PDF (Gemini Vision)
    INVOICE = "invoice"                # Invoice/receipt (Gemini Vision)
    FORM = "form"                      # Application form (Gemini Vision)
    PRODUCT_IMAGE = "product_image"    # Product photo (Vision Search)
    DAMAGE_IMAGE = "damage_image"      # Damage photo (Vision Search + note)
    WORD_DOC = "word_doc"              # DOCX/DOC
    SPREADSHEET = "spreadsheet"        # XLSX/XLS
    TEXT_FILE = "text_file"            # TXT/CSV
    VIDEO = "video"                    # Video file
    UNKNOWN = "unknown"


@dataclass
class ProcessedAttachment:
    """Result of processing an attachment"""
    filename: str
    attachment_type: AttachmentType
    content: str                       # Extracted text
    entities: Dict[str, Any]          # Structured data (model numbers, etc.)
    confidence: float                  # Processing confidence
    metadata: Dict[str, Any]          # Additional info
    processing_method: str            # Which processor was used
    processing_time: float
    error: Optional[str] = None


class EnhancedAttachmentProcessor:
    """
    Intelligent attachment processor that routes documents to optimal processing method.
    
    Cost optimization:
    - Simple PDFs: PyMuPDF (free)
    - Complex documents: Gemini Flash ($0.001-0.003 per image)
    - Detailed analysis: Gemini Pro ($0.05-0.10 per document)
    """
    
    def __init__(self):
        self.gemini_client = genai.Client(api_key=settings.gemini_api_key)
        # Import PDF processor lazily
        self._fitz = None
    
    def _get_pymupdf(self):
        """Lazy import of PyMuPDF"""
        if self._fitz is None:
            import fitz
            self._fitz = fitz
        return self._fitz
    
    def _download_attachment(self, url: str) -> Tuple[Optional[bytes], Optional[str]]:
        """Download attachment with Freshdesk authentication"""
        try:
            auth = HTTPBasicAuth(settings.freshdesk_api_key, "X")
            response = requests.get(url, auth=auth, timeout=30, stream=True)
            response.raise_for_status()
            return response.content, None
        except Exception as e:
            return None, f"Download failed: {str(e)}"
    
    def classify_attachment(
        self, 
        filename: str, 
        content_type: str,
        file_bytes: bytes = None
    ) -> AttachmentType:
        """
        Fast rule-based classification to route attachment to correct processor.
        
        Uses filename + content_type + optional quick peek at content.
        """
        fname_lower = filename.lower()
        
        # === IMAGE CLASSIFICATION ===
        if content_type.startswith("image/"):
            # Check if it's a product vs damage image
            damage_keywords = ["damage", "broken", "crack", "leak", "defect"]
            if any(kw in fname_lower for kw in damage_keywords):
                return AttachmentType.DAMAGE_IMAGE
            return AttachmentType.PRODUCT_IMAGE
        
        # === VIDEO ===
        if content_type.startswith("video/"):
            return AttachmentType.VIDEO
        
        # === PDF CLASSIFICATION ===
        if content_type in ("application/pdf", "application/x-pdf"):
            # Keywords suggest invoice/receipt
            invoice_keywords = ["invoice", "receipt", "bill", "order"]
            if any(kw in fname_lower for kw in invoice_keywords):
                return AttachmentType.INVOICE
            
            # Quick peek: check if PDF has text or is scanned
            if file_bytes:
                is_text_pdf = self._is_text_based_pdf(file_bytes)
                return AttachmentType.TEXT_PDF if is_text_pdf else AttachmentType.SCANNED_PDF
            
            # Default to text PDF (will fallback if needed)
            return AttachmentType.TEXT_PDF
        
        # === WORD DOCS ===
        if "word" in content_type or content_type.endswith(".document"):
            return AttachmentType.WORD_DOC
        
        # === SPREADSHEETS ===
        if "excel" in content_type or "spreadsheet" in content_type:
            return AttachmentType.SPREADSHEET
        
        # === TEXT FILES ===
        if content_type.startswith("text/"):
            return AttachmentType.TEXT_FILE
        
        return AttachmentType.UNKNOWN
    
    def _is_text_based_pdf(self, file_bytes: bytes) -> bool:
        """Quick check if PDF has extractable text (not scanned)"""
        try:
            fitz = self._get_pymupdf()
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            
            # Check first page for text
            if len(doc) > 0:
                text = doc[0].get_text()
                doc.close()
                # If we got substantial text, it's a text PDF
                return len(text.strip()) > 50
            
            doc.close()
            return False
        except:
            return False
    
    def process_attachment(
        self,
        attachment: Dict[str, Any]
    ) -> ProcessedAttachment:
        """
        Main entry point: Process any attachment intelligently.
        
        Routes to optimal processor based on type.
        """
        start_time = time.time()
        
        filename = attachment.get("name", "unknown")
        content_type = attachment.get("content_type", "")
        url = attachment.get("attachment_url")
        
        logger.info(f"🔍 Processing: {filename} ({content_type})")
        
        # Download file
        file_bytes, error = self._download_attachment(url)
        if error:
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.UNKNOWN,
                content="",
                entities={},
                confidence=0.0,
                metadata={"error": error},
                processing_method="download_failed",
                processing_time=time.time() - start_time,
                error=error
            )
        
        # Classify attachment type
        att_type = self.classify_attachment(filename, content_type, file_bytes)
        logger.info(f"📋 Classified as: {att_type.value}")
        
        # Route to appropriate processor
        if att_type == AttachmentType.TEXT_PDF:
            result = self._process_text_pdf(filename, file_bytes)
        
        elif att_type == AttachmentType.SCANNED_PDF:
            result = self._process_scanned_pdf(filename, file_bytes)
        
        elif att_type == AttachmentType.INVOICE:
            result = self._process_invoice(filename, file_bytes)
        
        elif att_type == AttachmentType.PRODUCT_IMAGE:
            result = self._process_product_image(filename, file_bytes, url)
        
        elif att_type == AttachmentType.DAMAGE_IMAGE:
            result = self._process_damage_image(filename, file_bytes, url)
        
        elif att_type == AttachmentType.WORD_DOC:
            result = self._process_word_doc(filename, file_bytes)
        
        elif att_type == AttachmentType.SPREADSHEET:
            result = self._process_spreadsheet(filename, file_bytes)
        
        elif att_type == AttachmentType.TEXT_FILE:
            result = self._process_text_file(filename, file_bytes)
        
        elif att_type == AttachmentType.VIDEO:
            result = self._process_video(filename, url)
        
        else:
            result = ProcessedAttachment(
                filename=filename,
                attachment_type=att_type,
                content=f"Unsupported file type: {content_type}",
                entities={},
                confidence=0.0,
                metadata={},
                processing_method="unsupported",
                processing_time=time.time() - start_time
            )
        
        result.processing_time = time.time() - start_time
        logger.info(f"✅ Processed in {result.processing_time:.2f}s using {result.processing_method}")
        
        return result
    
    # ==========================================
    # PROCESSOR METHODS
    # ==========================================
    
    def _process_text_pdf(self, filename: str, file_bytes: bytes) -> ProcessedAttachment:
        """Process text-based PDF using PyMuPDF (free, fast)"""
        try:
            fitz = self._get_pymupdf()
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            
            text_parts = []
            for page_num in range(min(len(doc), 50)):  # Max 50 pages
                text_parts.append(doc[page_num].get_text())
            
            doc.close()
            
            content = "\n\n".join(text_parts)
            
            # If text is empty, it might be scanned - fallback to Gemini
            if len(content.strip()) < 100:
                logger.warning("Text PDF extraction yielded little content, trying Gemini...")
                return self._process_scanned_pdf(filename, file_bytes)
            
            # Extract entities from text using simple patterns
            entities = self._extract_entities_regex(content)
            
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.TEXT_PDF,
                content=content,
                entities=entities,
                confidence=0.9,  # High confidence for native text
                metadata={"page_count": len(doc)},
                processing_method="pymupdf"
            )
        except Exception as e:
            logger.error(f"PyMuPDF failed: {e}, falling back to Gemini")
            return self._process_scanned_pdf(filename, file_bytes)
    
    def _process_scanned_pdf(self, filename: str, file_bytes: bytes) -> ProcessedAttachment:
        """Process scanned/image PDF using Gemini Vision OCR"""
        try:
            # Upload to Gemini
            file_obj = self.gemini_client.files.upload(
                file_path=io.BytesIO(file_bytes),
                mime_type="application/pdf"
            )
            
            # Use Gemini Flash for OCR (cheaper)
            prompt = """Extract ALL text from this PDF document.
            
Focus on:
- Product model numbers (e.g., HS6270MB, F2580CP)
- Order numbers (e.g., PO #12345)
- Part numbers
- Dates
- Quantities
- Product names

Return the extracted text preserving the document structure."""
            
            response = self.gemini_client.models.generate_content(
                model="gemini-1.5-flash",  # Cheap OCR
                contents=[
                    {"parts": [{"text": prompt}]},
                    {"parts": [{"file_data": {"file_uri": file_obj.uri, "mime_type": "application/pdf"}}]}
                ]
            )
            
            content = response.text or ""
            
            # Extract structured entities using Gemini
            entities = self._extract_entities_gemini(content)
            
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.SCANNED_PDF,
                content=content,
                entities=entities,
                confidence=0.85,
                metadata={"method": "gemini_flash_ocr"},
                processing_method="gemini_flash"
            )
        except Exception as e:
            logger.error(f"Gemini OCR failed: {e}")
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.SCANNED_PDF,
                content="",
                entities={},
                confidence=0.0,
                metadata={},
                processing_method="gemini_failed",
                error=str(e)
            )
    
    def _process_invoice(self, filename: str, file_bytes: bytes) -> ProcessedAttachment:
        """Process invoice/receipt using Gemini Pro for structured extraction"""
        try:
            # Upload to Gemini
            file_obj = self.gemini_client.files.upload(
                file_path=io.BytesIO(file_bytes),
                mime_type="application/pdf"
            )
            
            # Use Gemini Pro for complex invoice understanding
            prompt = """Analyze this invoice/receipt and extract:

1. Order Information:
   - Order number
   - Order date
   - Customer name

2. Product Details:
   - Product names
   - Model numbers
   - Quantities
   - Unit prices
   - Total price

3. Shipping/Delivery:
   - Shipping address
   - Expected delivery date

Respond with JSON:
{
  "order_number": "...",
  "order_date": "...",
  "customer_name": "...",
  "products": [
    {"name": "...", "model": "...", "quantity": X, "price": Y}
  ],
  "total": X.XX,
  "shipping_address": "...",
  "delivery_date": "..."
}"""
            
            response = self.gemini_client.models.generate_content(
                model="gemini-1.5-pro",  # Better for complex documents
                contents=[
                    {"parts": [{"text": prompt}]},
                    {"parts": [{"file_data": {"file_uri": file_obj.uri, "mime_type": "application/pdf"}}]}
                ],
                config=genai.types.GenerateContentConfig(
                    response_mime_type="application/json"
                )
            )
            
            # Parse JSON response
            import json
            entities = json.loads(response.text) if response.text else {}
            
            # Also get plain text summary
            content = self._format_invoice_summary(entities)
            
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.INVOICE,
                content=content,
                entities=entities,
                confidence=0.95,
                metadata={"structured_extraction": True},
                processing_method="gemini_pro"
            )
        except Exception as e:
            logger.error(f"Invoice processing failed: {e}")
            # Fallback to basic OCR
            return self._process_scanned_pdf(filename, file_bytes)
    
    def _process_product_image(self, filename: str, file_bytes: bytes, url: str) -> ProcessedAttachment:
        """Process product image - note for vision search tool"""
        return ProcessedAttachment(
            filename=filename,
            attachment_type=AttachmentType.PRODUCT_IMAGE,
            content=f"[Product image: {filename}]",
            entities={"image_url": url, "image_type": "product"},
            confidence=1.0,
            metadata={"for_vision_search": True},
            processing_method="image_handler"
        )
    
    def _process_damage_image(self, filename: str, file_bytes: bytes, url: str) -> ProcessedAttachment:
        """Process damage image - note for vision search + damage documentation"""
        return ProcessedAttachment(
            filename=filename,
            attachment_type=AttachmentType.DAMAGE_IMAGE,
            content=f"[Damage image: {filename}] - Customer reporting product damage",
            entities={"image_url": url, "image_type": "damage", "requires_visual_inspection": True},
            confidence=1.0,
            metadata={"for_vision_search": True, "damage_indicator": True},
            processing_method="image_handler"
        )
    
    def _process_word_doc(self, filename: str, file_bytes: bytes) -> ProcessedAttachment:
        """Process Word document"""
        try:
            from docx import Document
            doc = Document(io.BytesIO(file_bytes))
            
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            content = "\n\n".join(paragraphs)
            
            entities = self._extract_entities_regex(content)
            
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.WORD_DOC,
                content=content,
                entities=entities,
                confidence=0.9,
                metadata={},
                processing_method="python_docx"
            )
        except Exception as e:
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.WORD_DOC,
                content="",
                entities={},
                confidence=0.0,
                metadata={},
                processing_method="failed",
                error=str(e)
            )
    
    def _process_spreadsheet(self, filename: str, file_bytes: bytes) -> ProcessedAttachment:
        """Process Excel spreadsheet"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            
            sheets_content = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True, max_row=100):
                    row_text = " | ".join(str(cell) for cell in row if cell)
                    if row_text:
                        rows.append(row_text)
                if rows:
                    sheets_content.append(f"=== {sheet_name} ===\n" + "\n".join(rows))
            
            wb.close()
            content = "\n\n".join(sheets_content)
            entities = self._extract_entities_regex(content)
            
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.SPREADSHEET,
                content=content,
                entities=entities,
                confidence=0.9,
                metadata={},
                processing_method="openpyxl"
            )
        except Exception as e:
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.SPREADSHEET,
                content="",
                entities={},
                confidence=0.0,
                metadata={},
                processing_method="failed",
                error=str(e)
            )
    
    def _process_text_file(self, filename: str, file_bytes: bytes) -> ProcessedAttachment:
        """Process plain text file"""
        try:
            content = file_bytes.decode('utf-8', errors='replace')
            entities = self._extract_entities_regex(content)
            
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.TEXT_FILE,
                content=content,
                entities=entities,
                confidence=1.0,
                metadata={},
                processing_method="text_decode"
            )
        except Exception as e:
            return ProcessedAttachment(
                filename=filename,
                attachment_type=AttachmentType.TEXT_FILE,
                content="",
                entities={},
                confidence=0.0,
                metadata={},
                processing_method="failed",
                error=str(e)
            )
    
    def _process_video(self, filename: str, url: str) -> ProcessedAttachment:
        """Process video - placeholder for future frame extraction"""
        return ProcessedAttachment(
            filename=filename,
            attachment_type=AttachmentType.VIDEO,
            content=f"[Video file: {filename}] - Video processing not yet implemented",
            entities={"video_url": url, "requires_manual_review": True},
            confidence=0.5,
            metadata={"manual_review_needed": True},
            processing_method="video_placeholder"
        )
    
    # ==========================================
    # ENTITY EXTRACTION HELPERS
    # ==========================================
    
    def _extract_entities_regex(self, text: str) -> Dict[str, Any]:
        """Extract entities using regex patterns (fast, free)"""
        import re
        
        entities = {
            "model_numbers": [],
            "order_numbers": [],
            "part_numbers": [],
            "dates": [],
            "quantities": {}
        }
        
        # Model numbers (common patterns for plumbing products)
        model_patterns = [
            r'\b[A-Z]{2}\d{4}[A-Z]{2}\b',  # e.g., HS6270MB
            r'\b[A-Z]\d{4}[A-Z]{2}\b',     # e.g., F2580CP
            r'\b[A-Z]{1,2}-?\d{3,5}[A-Z]{0,2}\b'  # Various formats
        ]
        for pattern in model_patterns:
            entities["model_numbers"].extend(re.findall(pattern, text))
        
        # Order numbers
        order_patterns = [
            r'(?:PO|Order)\s*#?\s*(\d{4,})',
            r'Order\s+Number:\s*(\d{4,})'
        ]
        for pattern in order_patterns:
            entities["order_numbers"].extend(re.findall(pattern, text, re.IGNORECASE))
        
        # Part numbers
        part_patterns = [
            r'Part\s*#?\s*([A-Z0-9-]+)',
            r'Item\s*#?\s*([A-Z0-9-]+)'
        ]
        for pattern in part_patterns:
            entities["part_numbers"].extend(re.findall(pattern, text, re.IGNORECASE))
        
        # Dates
        date_patterns = [
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
            r'\d{4}-\d{2}-\d{2}'
        ]
        for pattern in date_patterns:
            entities["dates"].extend(re.findall(pattern, text))
        
        # Remove duplicates
        for key in entities:
            if isinstance(entities[key], list):
                entities[key] = list(set(entities[key]))
        
        return entities
    
    def _extract_entities_gemini(self, text: str) -> Dict[str, Any]:
        """Extract entities using Gemini (more accurate, costs ~$0.001)"""
        try:
            prompt = f"""Extract structured information from this text:

{text[:5000]}

Return JSON with:
{{
  "model_numbers": ["HS6270MB", ...],
  "order_numbers": ["12345", ...],
  "part_numbers": ["#ABC-123", ...],
  "product_names": ["Shower Head", ...],
  "quantities": {{"Shower Head": 2}},
  "dates": {{"order_date": "2024-01-15"}},
  "customer_info": {{"name": "...", "email": "..."}}
}}"""
            
            response = self.gemini_client.models.generate_content(
                model="gemini-1.5-flash",
                contents=prompt,
                config=genai.types.GenerateContentConfig(
                    response_mime_type="application/json",
                    temperature=0.1
                )
            )
            
            import json
            return json.loads(response.text) if response.text else {}
        except Exception as e:
            logger.warning(f"Gemini entity extraction failed: {e}, using regex fallback")
            return self._extract_entities_regex(text)
    
    def _format_invoice_summary(self, entities: Dict[str, Any]) -> str:
        """Format invoice entities into readable summary"""
        lines = [
            "=== INVOICE SUMMARY ===",
            f"Order: {entities.get('order_number', 'N/A')}",
            f"Date: {entities.get('order_date', 'N/A')}",
            f"Customer: {entities.get('customer_name', 'N/A')}",
            "",
            "Products:"
        ]
        
        for product in entities.get('products', []):
            lines.append(f"  - {product.get('name')} (Model: {product.get('model')}) x{product.get('quantity')}")
        
        lines.append(f"\nTotal: ${entities.get('total', 0):.2f}")
        
        return "\n".join(lines)


# ==========================================
# INTEGRATION WITH EXISTING SYSTEM
# ==========================================

def process_attachments_enhanced(attachments: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Drop-in replacement for process_all_attachments() with enhanced capabilities.
    
    This should replace the call in fetch_ticket.py
    """
    if not attachments:
        return {
            "extracted_content": "",
            "attachment_summary": [],
            "images": [],
            "stats": {"total": 0, "processed": 0, "failed": 0}
        }
    
    processor = EnhancedAttachmentProcessor()
    
    results: List[ProcessedAttachment] = []
    images: List[str] = []
    
    for att in attachments:
        # Skip images - handled by vision pipeline
        if att.get("content_type", "").startswith("image/"):
            images.append(att.get("attachment_url"))
            continue
        
        result = processor.process_attachment(att)
        results.append(result)
    
    # Build combined content
    content_parts = []
    attachment_summary = []
    
    for idx, result in enumerate(results, 1):
        summary = {
            "filename": result.filename,
            "type": result.attachment_type.value,
            "processing_method": result.processing_method,
            "confidence": result.confidence,
            "entities": result.entities,
            "chars": len(result.content),
            "error": result.error
        }
        attachment_summary.append(summary)
        
        if result.content:
            header = f"\n{'='*60}\n📄 {result.filename} ({result.attachment_type.value})\n{'='*60}\n"
            content_parts.append(header + result.content)
    
    combined_content = "\n".join(content_parts)
    
    return {
        "extracted_content": combined_content,
        "attachment_summary": attachment_summary,
        "images": images,
        "stats": {
            "total": len(attachments),
            "processed": len(results),
            "failed": sum(1 for r in results if r.error),
            "images": len(images)
        }
    }